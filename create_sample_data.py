"""
Creates the sample CRM Excel data file with all four sheets:
Customers, Interactions, Deals, Support Tickets
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_data():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: Customers ───
    ws_customers = wb.active
    ws_customers.title = "Customers"
    ws_customers.append([
        "Customer ID", "Company Name", "Industry", "Country", "City",
        "Account Owner", "Customer Tier", "Contract Start Date",
        "Contract End Date", "MRR (USD)", "ARR (USD)", "Health Score",
        "NPS Score", "Primary Contact Name", "Primary Contact Email"
    ])
    customers_data = [
        ["CUST-001", "Acme Corp", "Technology", "India", "Mumbai",
         "Rahu Sharma", "Gold", "2024-01-01", "2024-12-31",
         5000, 60000, 85, 8, "Priya Singh", "priya@acme.com"],
        ["CUST-002", "Beta Solutions", "Finance", "India", "Delhi",
         "Neha Gupta", "Silver", "2024-03-01", "2025-02-28",
         2000, 24000, 72, 7, "Arjun Mehta", "arjun@beta.com"],
        ["CUST-003", "Gamma Tech", "Healthcare", "India", "Bangalore",
         "Amit Verma", "Platinum", "2023-06-01", "2025-05-31",
         12000, 144000, 91, 9, "Kavya Rao", "kavya@gamma.com"],
        ["CUST-004", "Delta Corp", "Retail", "India", "Chennai",
         "Sanjay Jain", "Bronze", "2024-06-01", "2025-05-31",
         800, 9600, 45, 3, "Rahul Kumar", "rj@delta.com"],
        ["CUST-005", "Epsilon Ltd", "Education", "India", "Pune",
         "Rohit Sharma", "Silver", "2024-04-01", "2025-03-31",
         1500, 18000, 78, 7, "Anaya Bose", "anaya@epsilon.com"],
    ]
    for row in customers_data:
        ws_customers.append(row)

    # ─── Sheet 2: Interactions ───
    ws_interactions = wb.create_sheet("Interactions")
    ws_interactions.append([
        "Interaction ID", "Customer ID", "Date", "Type", "Subject",
        "Description", "Owner", "Duration (mins)", "Outcome",
        "Follow-up Date", "Follow-up Action"
    ])
    interactions_data = [
        ["INT-001", "CUST-001", "2024-11-15", "Call", "Quarterly Review",
         "Discussed Q4 goals and expansion plans", "Rahu Sharma", 45,
         "Positive", "2024-11-30", "Send expansion proposal"],
        ["INT-002", "CUST-002", "2024-11-18", "Email", "Renewal Discussion",
         "Sent renewal offer with 5% discount", "Neha Gupta", 0,
         "Pending", "2024-11-25", "Follow up by phone"],
        ["INT-003", "CUST-003", "2024-11-20", "Meeting", "Upsell Presentation",
         "Delivered analytics module demo", "Amit Verma", 60,
         "Positive", "2024-12-05", "Send contract for signature"],
        ["INT-004", "CUST-004", "2024-11-10", "Call", "Health Check",
         "Addressed support response time concerns", "Sanjay Jain", 30,
         "Neutral", "2024-11-20", "Escalate to Senior CSM"],
        ["INT-005", "CUST-005", "2024-11-22", "Meeting", "Onboarding Session 1",
         "Initial platform walkthrough", "Rohit Sharma", 90,
         "Positive", "2024-11-29", "Schedule onboarding session 2"],
    ]
    for row in interactions_data:
        ws_interactions.append(row)

    # ─── Sheet 3: Deals ───
    ws_deals = wb.create_sheet("Deals")
    ws_deals.append([
        "Deal ID", "Customer ID", "Deal Name", "Stage", "Value (USD)",
        "Expected Close Date", "Owner", "Probability (%)",
        "Product/Service", "Notes", "Created Date"
    ])
    deals_data = [
        ["DEAL-001", "CUST-001", "Enterprise Expansion", "Proposal", 25000,
         "2024-12-31", "Rahu Sharma", 65, "Pro Plan",
         "High potential - DM engaged", "2024-11-01"],
        ["DEAL-002", "CUST-002", "Annual Renewal", "Negotiation", 24000,
         "2025-02-28", "Neha Gupta", 80, "Standard Plan",
         "Discount requested", "2024-11-10"],
        ["DEAL-003", "CUST-003", "Analytics Module Add-on", "Demo", 18000,
         "2025-01-15", "Amit Verma", 50, "Analytics Module",
         "Needs IT approval", "2024-11-15"],
        ["DEAL-004", "CUST-005", "Initial Contract", "Closed Won", 18000,
         "2024-09-01", "Rohit Sharma", 100, "Standard Plan",
         "New customer win", "2024-09-01"],
        ["DEAL-005", "CUST-004", "Upgrade Attempt", "Closed Lost", 6000,
         "2024-10-15", "Sanjay Jain", 0, "Pro Plan",
         "Budget constraints", "2024-10-01"],
    ]
    for row in deals_data:
        ws_deals.append(row)

    # ─── Sheet 4: Support Tickets ───
    ws_tickets = wb.create_sheet("Support Tickets")
    ws_tickets.append([
        "Ticket ID", "Customer ID", "Created Date", "Priority", "Category",
        "Subject", "Description", "Assigned To", "Status",
        "Resolved Date", "Resolution Notes", "CSAT Score"
    ])
    tickets_data = [
        ["TKT-001", "CUST-001", "2024-11-10", "High", "Integration",
         "API not responding", "API calls timing out after 30s",
         "Dev Team", "Resolved", "2024-11-11",
         "Increased rate limit threshold", 5],
        ["TKT-002", "CUST-002", "2024-11-15", "Medium", "Billing",
         "Invoice amount mismatch", "Invoice shows $2200 instead of $2000",
         "Finance Team", "Open", "", "", ""],
        ["TKT-003", "CUST-003", "2024-11-18", "Low", "Feature Request",
         "Dark mode UI", "Request for dark theme across dashboard",
         "Product Team", "In Progress", "", "", ""],
        ["TKT-004", "CUST-004", "2024-11-05", "High", "Performance",
         "Dashboard loading slow", ">10 second load time",
         "Dev Team", "Resolved", "2024-11-07",
         "Optimized DB query", 4],
    ]
    for row in tickets_data:
        ws_tickets.append(row)

    # ─── Style headers ───
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ws in [ws_customers, ws_interactions, ws_deals, ws_tickets]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    wb.save("CRM_Sample_Data_Template.xlsx")
    print("✅ CRM_Sample_Data_Template.xlsx created successfully!")

if __name__ == "__main__":
    create_sample_data()
