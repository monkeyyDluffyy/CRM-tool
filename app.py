"""
CRM Tool Backend - Flask API with Excel as data source
Provides full CRUD operations on the Excel file with live-reload support.
"""

import os
import sys
import json
import time
import threading
import copy
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# ─── EXCEL FILE PATH CONFIGURATION ───
# You can set the path to your Excel file in 3 ways (in order of priority):
#
#   1. Command-line argument:   python3 app.py /path/to/your/file.xlsx
#   2. Environment variable:    export CRM_EXCEL_FILE=/path/to/your/file.xlsx
#   3. Default:                 CRM_Sample_Data_Template.xlsx (in this folder)
#
DEFAULT_EXCEL = "/home/abhay/Downloads/CRM_Sample_Data_Template.xlsx"

if len(sys.argv) > 1 and sys.argv[1].endswith('.xlsx'):
    EXCEL_FILE = os.path.abspath(sys.argv[1])
elif os.environ.get('CRM_EXCEL_FILE'):
    EXCEL_FILE = os.path.abspath(os.environ['CRM_EXCEL_FILE'])
else:
    EXCEL_FILE = DEFAULT_EXCEL

# Track file modification for live-reload
file_lock = threading.Lock()


def read_excel_data():
    """Read all sheets from the Excel file. Always reads fresh from disk."""
    with file_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        except Exception as e:
            print(f"[ERROR] Cannot read Excel file: {e}")
            return {}

        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sheet_name] = {'headers': [], 'data': []}
                continue

            headers = [str(h) if h else f'Column_{i}' for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                # Skip the instruction template row if present
                first_val = str(row[0]).strip() if row and len(row) > 0 and row[0] is not None else ""
                if first_val.startswith("AUTO:"):
                    continue

                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        if isinstance(val, datetime):
                            val = val.strftime('%Y-%m-%d')
                        elif val is None:
                            val = ''
                        row_dict[headers[i]] = val
                data.append(row_dict)
            result[sheet_name] = {'headers': headers, 'data': data}

        wb.close()
        return result


def write_to_excel(sheet_name, data_rows, headers):
    """Write data back to a specific sheet in the Excel file."""
    with file_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except Exception:
            wb = openpyxl.Workbook()

        # Remove existing sheet if it exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        # Create sheet at correct position
        desired_order = ['Customers', 'Interactions', 'Deals', 'Support Tickets']
        if sheet_name in desired_order:
            idx = desired_order.index(sheet_name)
            # Clamp to valid range
            idx = min(idx, len(wb.sheetnames))
            ws = wb.create_sheet(sheet_name, idx)
        else:
            ws = wb.create_sheet(sheet_name)

        # Write headers
        ws.append(headers)

        # Write data rows
        for row_dict in data_rows:
            row = []
            for h in headers:
                val = row_dict.get(h, '')
                # Try to convert numeric strings
                if isinstance(val, str) and val.strip():
                    try:
                        if '.' in val:
                            val = float(val)
                        else:
                            val = int(val)
                    except (ValueError, TypeError):
                        pass
                row.append(val)
            ws.append(row)

        # Remove default sheet if it exists and is empty
        if 'Sheet' in wb.sheetnames and len(wb['Sheet'].dimensions) <= 2:
            del wb['Sheet']

        wb.save(EXCEL_FILE)
        wb.close()
        print(f"[SYNC] ✅ Excel file updated: {sheet_name} ({len(data_rows)} rows)")


# ─── API Routes ───

@app.route('/')
def serve_index():
    return send_from_directory('static', 'index.html')


@app.route('/api/data', methods=['GET'])
def get_all_data():
    """Get all data from all sheets — always reads fresh from Excel."""
    data = read_excel_data()
    return jsonify(data)


@app.route('/api/data/<sheet_name>', methods=['GET'])
def get_sheet_data(sheet_name):
    """Get data from a specific sheet."""
    data = read_excel_data()
    if sheet_name in data:
        return jsonify(data[sheet_name])
    return jsonify({'error': f'Sheet "{sheet_name}" not found'}), 404


@app.route('/api/data/<sheet_name>', methods=['POST'])
def add_row(sheet_name):
    """Add a new row to a specific sheet."""
    data = read_excel_data()
    if sheet_name not in data:
        return jsonify({'error': f'Sheet "{sheet_name}" not found'}), 404

    new_row = request.json
    if not new_row:
        return jsonify({'error': 'No data provided'}), 400

    headers = data[sheet_name]['headers']
    rows = data[sheet_name]['data']
    rows.append(new_row)

    write_to_excel(sheet_name, rows, headers)
    return jsonify({'success': True, 'message': 'Row added successfully'})


@app.route('/api/data/<sheet_name>/<int:row_index>', methods=['PUT'])
def update_row(sheet_name, row_index):
    """Update a specific row in a sheet."""
    data = read_excel_data()
    if sheet_name not in data:
        return jsonify({'error': f'Sheet "{sheet_name}" not found'}), 404

    rows = data[sheet_name]['data']
    if row_index < 0 or row_index >= len(rows):
        return jsonify({'error': 'Row index out of range'}), 400

    updated_row = request.json
    if not updated_row:
        return jsonify({'error': 'No data provided'}), 400

    headers = data[sheet_name]['headers']
    for key, value in updated_row.items():
        if key in headers:
            rows[row_index][key] = value

    write_to_excel(sheet_name, rows, headers)
    return jsonify({'success': True, 'message': 'Row updated successfully'})


@app.route('/api/data/<sheet_name>/<int:row_index>', methods=['DELETE'])
def delete_row(sheet_name, row_index):
    """Delete a specific row from a sheet."""
    data = read_excel_data()
    if sheet_name not in data:
        return jsonify({'error': f'Sheet "{sheet_name}" not found'}), 404

    rows = data[sheet_name]['data']
    if row_index < 0 or row_index >= len(rows):
        return jsonify({'error': 'Row index out of range'}), 400

    headers = data[sheet_name]['headers']
    deleted = rows.pop(row_index)
    print(f"[DELETE] Removed row {row_index} from {sheet_name}: {deleted}")

    write_to_excel(sheet_name, rows, headers)
    return jsonify({'success': True, 'message': 'Row deleted successfully'})


@app.route('/api/check-update', methods=['GET'])
def check_update():
    """Check if Excel file has been modified (for polling)."""
    try:
        current_mtime = os.path.getmtime(EXCEL_FILE)
    except OSError:
        current_mtime = 0
    return jsonify({
        'modified': True,  # Always report modified so frontend always fetches fresh
        'timestamp': current_mtime
    })


def safe_float(val):
    try:
        if val is None or str(val).strip() == '': return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    """Get aggregated dashboard stats."""
    data = read_excel_data()

    customers = data.get('Customers', {}).get('data', [])
    deals = data.get('Deals', {}).get('data', [])
    tickets = data.get('Support Tickets', {}).get('data', [])
    interactions = data.get('Interactions', {}).get('data', [])

    # Customer stats
    total_customers = len(customers)
    total_mrr = sum(safe_float(c.get('MRR (USD)', 0)) for c in customers)
    total_arr = sum(safe_float(c.get('ARR (USD)', 0)) for c in customers)
    avg_health = sum(safe_float(c.get('Health Score', 0)) for c in customers) / max(total_customers, 1)
    avg_nps = sum(safe_float(c.get('NPS Score', 0)) for c in customers) / max(total_customers, 1)

    # Tier distribution
    tier_dist = {}
    for c in customers:
        tier = c.get('Customer Tier', 'Unknown')
        tier_dist[tier] = tier_dist.get(tier, 0) + 1

    # Deal stats
    total_pipeline = sum(
        safe_float(d.get('Value (USD)', 0))
        for d in deals
        if d.get('Stage', '') not in ('Closed Won', 'Closed Lost')
    )
    won_deals = sum(1 for d in deals if d.get('Stage') == 'Closed Won')
    lost_deals = sum(1 for d in deals if d.get('Stage') == 'Closed Lost')
    active_deals = sum(1 for d in deals if d.get('Stage') not in ('Closed Won', 'Closed Lost'))

    # Deal stage distribution
    stage_dist = {}
    for d in deals:
        stage = d.get('Stage', 'Unknown')
        stage_dist[stage] = stage_dist.get(stage, 0) + 1

    # Ticket stats
    open_tickets = sum(1 for t in tickets if t.get('Status') in ('Open', 'In Progress'))
    resolved_tickets = sum(1 for t in tickets if t.get('Status') == 'Resolved')
    high_priority = sum(1 for t in tickets if t.get('Priority') == 'High')

    # Priority distribution
    priority_dist = {}
    for t in tickets:
        p = t.get('Priority', 'Unknown')
        priority_dist[p] = priority_dist.get(p, 0) + 1

    # Interaction stats
    total_interactions = len(interactions)
    positive_outcomes = sum(1 for i in interactions if i.get('Outcome') == 'Positive')

    # Industry distribution
    industry_dist = {}
    for c in customers:
        ind = c.get('Industry', 'Unknown')
        industry_dist[ind] = industry_dist.get(ind, 0) + 1

    # MRR by tier
    mrr_by_tier = {}
    for c in customers:
        tier = c.get('Customer Tier', 'Unknown')
        mrr = safe_float(c.get('MRR (USD)', 0))
        mrr_by_tier[tier] = mrr_by_tier.get(tier, 0) + mrr

    return jsonify({
        'customers': {
            'total': total_customers,
            'total_mrr': total_mrr,
            'total_arr': total_arr,
            'avg_health': round(avg_health, 1),
            'avg_nps': round(avg_nps, 1),
            'tier_distribution': tier_dist,
            'industry_distribution': industry_dist,
            'mrr_by_tier': mrr_by_tier,
        },
        'deals': {
            'total_pipeline': total_pipeline,
            'won': won_deals,
            'lost': lost_deals,
            'active': active_deals,
            'stage_distribution': stage_dist,
        },
        'tickets': {
            'open': open_tickets,
            'resolved': resolved_tickets,
            'high_priority': high_priority,
            'priority_distribution': priority_dist,
        },
        'interactions': {
            'total': total_interactions,
            'positive': positive_outcomes,
        }
    })


if __name__ == '__main__':
    print(f"\n🚀 CRM Tool running at http://localhost:5000")
    print(f"📊 Excel data source: {EXCEL_FILE}")
    print(f"🔄 Auto-sync: reads fresh from Excel on every request\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
